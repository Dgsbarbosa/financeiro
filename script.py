#!/usr/bin/env python3

# -*- coding: utf-8 -*-

import calendar
from datetime import datetime
import locale
import openpyxl
from openpyxl.worksheet.header_footer import HeaderFooter
from openpyxl.styles import Border, Side
import re
from openpyxl.styles import Alignment, NamedStyle
import csv
import time
import tkinter as tk
from tkinter import filedialog
import os
import sys
import tqdm
from tqdm import tqdm
import pandas as pd

# Definir a localizacao para portugues do Brasil
locale.setlocale(locale.LC_ALL, "pt_BR.utf-8")
# Fixar o domingo como primeiro dia
calendar.setfirstweekday(calendar.SUNDAY)

#caminho global arquivo de despesas e 
arquivo_de_entradas = ""
arquivo_de_despesas = ""



def main():
    
    print()
    barra_de_progresso("Preparando dados...")
    time.sleep(0.5)

    criacao_planilha()
    try:
        ...
    except Exception as e:
        print(f"\nHouve um erro na analise dos dados do programa:\n{e}")
        print("Verifique se os arquivos de entrada e saida esta na mesma pasta do programa")
        raise SystemExit

def excel_para_csv():

    try:
        global arquivo_de_entradas
        global arquivo_de_despesas
        #Leitura da pasta de trabalho
        pasta_excel = pd.ExcelFile("Contas Basicas.xlsx")
        
        barra_de_progresso(f'Abrindo o arquivo "Contas Basicas"')
        
        time.sleep(1)
        
        # Leitura da folha 'entrada' e salvamento como CSV
        entrada_df = pasta_excel.parse('Recebimentos')
        entrada_df['Vencimento'] = entrada_df['Vencimento'].map('{:.0f}'.format)
        entrada_csv_path = f"entradas.csv"
        entrada_df.to_csv(entrada_csv_path, index=False)

        # Leitura da folha 'despesa' e salvamento como CSV
        despesa_df = pasta_excel.parse('Despesas')
        despesa_df['Vencimento'] = despesa_df['Vencimento'].map('{:.0f}'.format)
        despesa_csv_path = f"despesas.csv"
        despesa_df.to_csv(despesa_csv_path, index=False)
        time.sleep(1)
        print("\nConvertendo pasta de trabalho para arquivos CSV")
        
        arquivo_de_entradas = entrada_csv_path
        arquivo_de_despesas = despesa_csv_path
        barra_de_progresso('Progresso ')
        
    except Exception as e:
        
        print(f'\nHouve um erro na abertura do arquivo "Contas Basicas"\n{e}')
        print('\nVerifique se o arquivo "Contas Basicas" esta na mesma pasta do programa')
        sys.exit()

def barra_de_progresso(texto_descricao):
    for i in tqdm(range(5),desc=texto_descricao,bar_format="{desc} {bar} {percentage}%", ncols=80):
        time.sleep(0.2)
    
def verificar_vencimentos():
    global arquivo_de_entradas 
    global arquivo_de_despesas
    
    print("\nOs vencimentos das contas basicas nao podem ter o mesmo dia de vencimento...")

    time.sleep(2)

    print("\nCaso haja vencimento duplicado sera solicitado a mudanca\n")
    
    time.sleep(2)
    
    # barra de progresso
    barra_de_progresso("Analizando vencimentos ")    

    lista_de_contas_original = []
    
    nova_lista_contas = []
    
    
    # Le o arquivo de despesa
    with open(arquivo_de_despesas, "r", newline="") as csvfile:
        leitor_contas = csv.DictReader(csvfile)
        
        # Percorre o csv de despesas
        for conta in leitor_contas:
            
            if not conta['Vencimento'].isdigit():
                conta['Vencimento'] = 0
            lista_de_contas_original.append(conta)
    
    
    lista_de_contas_copy_1 = lista_de_contas_original.copy()
    lista_de_contas_copy_2 = lista_de_contas_original.copy()
    
    # percorre a lista de contas originais
    for conta in lista_de_contas_copy_2:
        
        nome_conta = conta['Nome']
        
        vencimento_conta = int(conta['Vencimento'])
        

        vencimento_duplicado = []
        vencimento_duplicado.append(conta)
        # percorre a copia da lista de contas
        for conta_copy in lista_de_contas_copy_1:
            
            nome_conta_copy = conta_copy['Nome']  
                                    
            vencimento_conta_copy = int(conta_copy['Vencimento'])
            
            nomes_iguais = nome_conta == nome_conta_copy
            venc_iguais = vencimento_conta == vencimento_conta_copy 
            
            # Se os nomes forem diferentes e os vencimentos forem iguais coloca na lista de mesmo vencimento e retira da copia da lista de contas 
            if nomes_iguais == False and venc_iguais == True:
                vencimento_duplicado.append(conta_copy)
                
                lista_de_contas_copy_2.remove(conta_copy)
                
        # checa as contas iguais
        while len(vencimento_duplicado) != 0: 
            if len(vencimento_duplicado) == 1:
                nova_lista_contas.append(vencimento_duplicado[0])
                lista_de_contas_copy_1.remove(vencimento_duplicado[0])
                vencimento_duplicado = []        
                
            elif len(vencimento_duplicado) > 1:
                qtd_vencimento_duplicado = len(vencimento_duplicado)
                
                # Verifica se o dia eh 0
                dia_vencimento = vencimento_duplicado[0]

                if dia_vencimento['Vencimento'] == 0:
                    
                    print(f"\nForam encontradas {qtd_vencimento_duplicado} contas com vencimento no dia 0...")
                    
                    time.sleep(2.5)
                    
                    print(f"\nEssas contas serao serao adicionadas em dias que nao ha despesa anual ou mensal")
                    
                    time.sleep(2.5)

                    for venc_zero in vencimento_duplicado:
                        venc_zero_nome = venc_zero['Nome'].capitalize()
                        venc_zero_valor = venc_zero['Valor']
                        venc_zero_tipo = venc_zero['Tipo'].capitalize()
                        
                        print(f'\nA Conta: {venc_zero_nome} foi encontrada')

                        barra_de_progresso("Adicionando conta ")
                        
                        nova_lista_contas.append(venc_zero)
                        
                        print(f'A Conta: {venc_zero_nome}, Valor: ${venc_zero_valor}, Tipo: {venc_zero_tipo}, foi adicionada')

                        time.sleep(1)
                    vencimento_duplicado = []
                    continue
                    
                print(f'\nForam encontradas {qtd_vencimento_duplicado} com vencimento no dia {vencimento_duplicado[0]["Vencimento"]}.')
                print("Necessario alterar a data de uma delas\n")

                
                max_len_nome = max(len(conta['Nome'].capitalize()) for conta in vencimento_duplicado)
                max_len_valor = max(len(f"Valor: ${conta['Valor']}") for conta in vencimento_duplicado)
                max_len_tipo = max(len(conta['Tipo'].capitalize()) for conta in vencimento_duplicado)
                
                # Gera as opcoes da lista mesmo vencimento
                for i, conta in enumerate(vencimento_duplicado, start=1):
                    nome_conta = conta['Nome'].upper()
                    valor_conta = f"Valor: ${conta['Valor']}"
                    tipo_conta = conta['Tipo'].upper()

                    print(f"{i}. Conta: {nome_conta.ljust(max_len_nome)}, {valor_conta.ljust(max_len_valor)}, Tipo: {tipo_conta.ljust(max_len_tipo)}")      
                             
                print("0. Sair do programa")
            
                #Escolhe a conta que vai ser alterada                
                while True:
                    opcao_escolhida = input(f"\nEscolha uma opcao entre 1 e {qtd_vencimento_duplicado}: ")
                    
                    if opcao_escolhida.isdigit() and 1 <= int(opcao_escolhida) <= qtd_vencimento_duplicado:
                        opcao_escolhida = int(opcao_escolhida)    
                        break
                    elif opcao_escolhida == "0":
                        sys.exit()
                    else:
                        print("Opcao invalida")
                        continue
                
                index_opcao = opcao_escolhida -1
                conta_opcao = vencimento_duplicado[index_opcao]             
                
                # imprime a conta escolhida e o vencimento atual
                print(f"\nOpcao escolhida\n{opcao_escolhida}. Nome: {conta_opcao['Nome'].upper()}, Valor: ${conta_opcao['Valor']}, Tipo: {conta_opcao['Tipo'].upper()}")
                print(f"Vencimento atual: dia {conta_opcao['Vencimento']}")
                
                # escolhe o novo vencimento e checa se ja tem o vencimento nas listas
                novo_vencimento_valido = False                
                while novo_vencimento_valido == False:
                    
                    novo_vencimento_valido = False
                    novo_vencimento = input(f"\nEscollha o novo dia de vencimento (dd): ")


                    if not novo_vencimento.isdigit():
                        print("\nO novo vencimento deve ser um numero")
                        continue

                        
                    novo_vencimento = int(novo_vencimento)
                    
                    if novo_vencimento < 1 or novo_vencimento > 31:
                        print("\nEscolha um numero de 1 a 31")
                        continue
                    
                    # Barra de progressao
                    print()
                    barra_de_progresso("Analisando dia do novo vencimento")

                    # verifica se o novo vencimento eh diferente da copia da lista de contas e da nova lista de contas 

                    venc_copy_valido = True
                    for item in lista_de_contas_copy_1:
                        if novo_vencimento == int(item['Vencimento']):                        
                            print(f"\nHa uma conta nao analisada com vencimento no dia: {novo_vencimento}")
                            print(f"Conta: {item['Nome'].upper()}")
                            venc_copy_valido = False
                            break
                            
                    

                    venc_nova_lista_valido = True
                    for item in nova_lista_contas:
                        if novo_vencimento == int(item['Vencimento']):                        
                            print(f"\nHa uma conta ja analisada com vencimento no dia: {novo_vencimento}")
                            print(f"Conta: {item['Nome'].upper()}")
                            venc_nova_lista_valido = False
                            break
                    
                    
                    if venc_copy_valido == True and venc_nova_lista_valido == True:
                        
                        conta_opcao['Vencimento'] = str(novo_vencimento) 
                        novo_vencimento_valido = True
                    
                
                
                # Mostra a opcao escolhida e o novo vencimento, salva na nova lista e remove da lista mesmo vencimento
                print(f"\n{opcao_escolhida}. Nome: {conta_opcao['Nome'].upper()}, Valor: ${conta_opcao['Valor']}, Tipo: {conta_opcao['Tipo']}")
                    
                print(f"Novo Vencimento: dia {conta_opcao['Vencimento']}\n")
                
                
                # Barra de progresso
                barra_de_progresso(f"Alterando vencimento de {conta_opcao['Nome'].upper()}")

                # Salva a conta escolhida com novo vencimento
                nova_lista_contas.append(conta_opcao)
    
                # Remove a conta escolhida da lista de mesmo vencimento
                vencimento_duplicado.remove(conta_opcao)
                
                #Remove a conta escolhida da copia 1 de lista 
                lista_de_contas_copy_1.remove(conta_opcao)
                
                time.sleep(0.5)
                print(f"\nA conta {conta_opcao['Nome'].upper()} foi alterada para vencimento dia {conta_opcao['Vencimento']}")
                time.sleep(2)
        
    # Escrevendo a nova lista no arquivo CSV
    with open(arquivo_de_despesas, "w", newline="") as csvfile:
        escritor_csv = csv.DictWriter(csvfile, fieldnames=nova_lista_contas[0].keys())
        
        # Escrevendo o cabecalho
        escritor_csv.writeheader()
        
        # Escrevendo os dados
        escritor_csv.writerows(nova_lista_contas)
    
    print()
    barra_de_progresso("Atualizando lista de despesas")

def exibir_calendario_por_semana(ano, mes):
    # Obter o calendario do mes como uma lista de listas
    calendario_mes = calendar.monthcalendar(ano, mes)

    nome_do_mes = calendar.month_name[mes]

    weeks = []

    # Iterar sobre as semanas e exibir os dias
    for i, semana in enumerate(calendario_mes, start=1):
        numero_da_semana = f"Semana {i}"

        week = {}
        week[i] = []

        for dia in semana:
            if dia == 0:
                ...
            else:
                data_completa = datetime(ano, mes, dia)

                dia_da_semana_str = data_completa.strftime("%A")
                mes_str = data_completa.strftime("%B")
                ano_str = data_completa.strftime("%Y")

                data_completa_str = f"{dia_da_semana_str.capitalize()}, {dia:2} de {mes_str} de {ano_str}"

                week[i].append(data_completa_str)

        # print(week)
        weeks.append(week)

    # for w in weeks:
    #     print(w)

    return weeks

def calendario_anual():
    ano = datetime.now().year

    # print("ano", ano)

    calendario = {}
    for i in range(1, 13):
        nome_do_mes = calendar.month_name[i]
        calendario_semanal = exibir_calendario_por_semana(ano, i)

        calendario[nome_do_mes] = calendario_semanal

    return calendario

def criacao_planilha():
    calendario_completo = calendario_anual()
    wb = openpyxl.Workbook()

    # Defina um estilo para formatar valores em dinheiro
    money_style = NamedStyle(
        name="br_number_style", number_format='_("R$ "* #,##0.00_)'
    )

    # Crie um objeto de borda
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    lista_contas_anuais = contas_anuais()

    
    for mes, semanas in calendario_completo.items():
        ws = wb.create_sheet(title=mes.capitalize())
        ano_atual = datetime.now().year

        ws.append(
            [
                f"Semanas {ano_atual}",
                "Descricao",
                "Recebimentos ",
                "Contas da casa",
                "Valor a pagar",
            ]
        )

        ws.append([])

        for semana in semanas:
            for numero_semana, dias_semana_str in semana.items():
                contas_semana = contas_semanais()

                ws.append([f"Semana {numero_semana}"])

                for dia_semana in dias_semana_str:
                    dia_int = extrair_dia(dia_semana)

                    dia_da_semana_split = dia_semana.split(" ")
                    dia_da_semana_split = dia_da_semana_split[:-2]
                    dia_da_semana_split = " ".join(dia_da_semana_split)

                    dia_semana = dia_da_semana_split

                    for conta_anual in lista_contas_anuais:
                        
                        
                        if int(conta_anual["Vencimento"]) == dia_int:
                            padrao = re.compile(r"\bmar.*o\b", re.IGNORECASE)

                            if padrao.search(mes):
                                mes = "marco"

                            lista_de_parcelas = conta_anual["parcelas"]

                            if mes in lista_de_parcelas:
                                indice_parcela = lista_de_parcelas.index(mes)

                                numero_parcela = indice_parcela + 1

                                qtd_parcelas = conta_anual["qtd_parcela"]

                                nome_conta_alterado = f"{conta_anual['Nome'].capitalize()} {numero_parcela}/{qtd_parcelas}"

                                ws.append(
                                    [
                                        dia_semana,
                                        "",
                                        "",
                                        nome_conta_alterado,
                                        conta_anual["Valor"],
                                    ]
                                )

                                ws["E{}".format(ws.max_row)].style = money_style

                    # Entradas e saidas mensais e semanais
                    conta_mensal = contas_mensais(dia_int)
                    
                    # print("conta",conta_mensal)
                    
                    entrada_mensal = entradas(dia_int)

                    if entrada_mensal:
                        nome_entrada = entrada_mensal["Nome"].capitalize()
                        valor_entrada = entrada_mensal["Valor"]

                    else:
                        nome_entrada = ""
                        valor_entrada = ""

                    if conta_mensal:
                        nome_conta = conta_mensal["Nome"].capitalize()
                        valor_conta = conta_mensal["Valor"]

                    elif contas_semana:
                        conta_semana = contas_semana[0]
                        nome_a_excluir = conta_semana["Nome"]

                        nova_lista_contas = [
                            conta
                            for conta in contas_semana
                            if conta["Nome"] != nome_a_excluir
                        ]
                        contas_semana = nova_lista_contas

                        nome_conta = conta_semana["Nome"]
                        valor_conta = conta_semana["Valor"]

                    else:
                        nome_conta = ""
                        valor_conta = ""

                    ws.append(
                        [
                            dia_semana,
                            nome_entrada,
                            valor_entrada,
                            nome_conta,
                            valor_conta,
                        ]
                    )

                    ws["C{}".format(ws.max_row)].style = money_style

                    ws["E{}".format(ws.max_row)].style = money_style

                # Adicione uma linha vazia apÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â³s cada semana

                ws.append([])

        # Adicione uma linha vazia apÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â³s cada mÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Âªs
        ws.append([])
        ws.append([])

        # Ajuste a largura da coluna
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Aplique a borda a todas as cÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â©lulas
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Remova a planilha padrao criada e salve o arquivo
    del wb["Sheet"]

    try:
        print()
        barra_de_progresso("Gerando planilha....")

        print("\nPlanilha gerada com sucesso\n")
        
        time.sleep(1)
        
        while True:
            salvar_planilha = input("Gostaria de salvar a planilha (s/n): ").lower()
            
            time.sleep(0.5)
        
            if salvar_planilha == "s":
                
                print("\nEscolha o local para salvar\n")
                
                path = salvar_arquivo()

                time.sleep(0.5)
                wb.save(path)
                
                barra_de_progresso("Salvando planilha")
                print("\nPlanilha salva com sucesso")
                print(f"O arquivo foi salvo em {path}\n")
                
                break
            elif salvar_planilha == "n":
                print("\nO programa ira fechar e a planilha nao sera salva")
                while True:
                    fechar_programa = input("\nQuer encerrar o programa (s/n): \n").lower()
                    if fechar_programa == "s":
                        sys.exit()
                    elif fechar_programa == "n":
                        
                        break
                    else:
                        print("\nOpcao invalida, escolha (s/n)\n")
                        continue
            
            else:
                print("Opcao invalida, escolha (s/n)")
    except Exception as e:
        print("Houve um erro na criacao da planilha")
        print(f"Erro {e}")

def extrair_dia(data_completa):
    # Expressao regular para extrair o dia e mes

    dia_pattern = r"\b\d{1,2}\b"

    # Encontrar todos os dias nas strings
    dia = re.findall(dia_pattern, data_completa)
    dia = dia[0]
    dia = int(dia)

    return dia

def contas_mensais(dia):
    contas_lidas = ler_contas_csv()
    
    
    # Exibindo os resultados
    for conta in contas_lidas:
        vencimento = int(conta["Vencimento"])
        if vencimento == dia and conta["Tipo"] == "mensal":
            
            
            return conta

def contas_semanais():
    contas_lida = ler_contas_csv()

    contas_semana = []

    for conta in contas_lida:
        if conta["Tipo"] == "semanal":
            contas_semana.append(conta)

    return contas_semana

def contas_anuais():
    meses = {
        1: "janeiro",
        2: "fevereiro",
        3: "marco",
        4: "abril",
        5: "maio",
        6: "junho",
        7: "julho",
        8: "agosto",
        9: "setembro",
        10: "outubro",
        11: "novembro",
        12: "dezembro",
    }

    contas_lidas = ler_contas_csv()
    
    lista_de_contas_anuais = []

    # Percorre lista de contas procurando contas anuais
    for conta in contas_lidas:
        if conta["Tipo"] == "anual":
            nome_da_conta = conta["Nome"].capitalize()


            print(f"\nUma conta anual foi encontrada: \n{nome_da_conta}")

            while True:
                parcelado = input(
                    f"\nA conta: {nome_da_conta} sera parcelada? (s/n): "
                ).lower()

                if parcelado == "s" or parcelado == "n":
                    break
                else:
                    print(
                        "Opcao invalida. Por favor, escolha 's' para sim ou 'n' para nao."
                    )

            # Se o valor for parcelado
            if parcelado == "s":
                controle_while = False

                while controle_while == False:
                    print()
                    for mes in meses:
                        print(f"({mes}-{meses[mes].capitalize()})", end=" ")

                    print("\n\nMes da primeira parcela")

                    mes_inicio = input("\nDigite um numero de 1 a 12: ")

                    if mes_inicio.isdigit() and 1 <= int(mes_inicio) <= 12:
                        mes_inicio = int(mes_inicio)
                        while True:
                            qtd_parcela = input("Quantidade de parcelas: ")

                            if qtd_parcela.isdigit() and 1 <= int(qtd_parcela) <= 12:
                                qtd_parcela = int(qtd_parcela)
                                mes_final = mes_inicio + (qtd_parcela - 1)

                                if mes_final <= 12:
                                    controle_while = True
                                    break

                                else:
                                    print(
                                        "\nO inicio das parcelas e a quantidade de parcelas excedem os meses do ano"
                                    )
                                    break

                            else:
                                print("\nDigite um numero de 1 a 12")

                    else:
                        print("Digite um numero de 1 a 12")

                meses_parcela = []

                for i in range(mes_inicio, mes_final + 1):
                    meses_parcela.append(meses[i])

                valor_total = conta["Valor"]
                conta["Valor"] = dividir_parcelas(valor_total, qtd_parcela)

                conta["parcelas"] = meses_parcela

                conta["qtd_parcela"] = qtd_parcela

                lista_de_contas_anuais.append(conta)

                barra_de_progresso("Calculando parcelas...")

            # Caso a conta anual nao seja parcelada
            else:
                
                for mes in meses:
                    print(f"({mes}-{meses[mes].capitalize()})", end=" ")
                print("\nQual o mes de vencimento\n")


                while True:
                    mes_vencimento = input("\nDigite um numero de 1 a 12: ")

                    if mes_vencimento.isdigit() and 1 <= int(mes_vencimento) <= 12:
                        mes_vencimento = int(mes_vencimento)
                        break

                conta["parcelas"] = [meses[mes_vencimento]]
                conta["qtd_parcela"] = 1

                lista_de_contas_anuais.append(conta)
                barra_de_progresso(f"Inserindo conta {nome_da_conta}...")
                
          
    return lista_de_contas_anuais

# Divisao de parcelas das contas anuais
def dividir_parcelas(valor, qtd_parcelas):
    valor_da_parcela = valor / qtd_parcelas

    # print(f"dividir_parcelas - valor: {valor}, qtd_parcelas: {qtd_parcelas} = valor da parcela {valor_da_parcela}")
    return valor_da_parcela

def entradas(dia):
    caminho_arquivo = arquivo_de_entradas

    with open(caminho_arquivo, "r", newline="") as csvfile:
        leitor_csv = csv.DictReader(csvfile)

        for linha in leitor_csv:
            if int(linha["Vencimento"]) == dia:
                linha["Valor"] = float(linha["Valor"])
                return linha

# Ler arquivo e retornar contas
def ler_contas_csv():
    contas = []
    caminho_arquivo = arquivo_de_despesas

    with open(caminho_arquivo, "r", newline="") as csvfile:
        leitor_csv = csv.DictReader(csvfile)

        for linha in leitor_csv:
 
            contas.append(
                {
                    "Nome": linha["Nome"],
                    "Vencimento": linha["Vencimento"],
                    "Valor": float(linha["Valor"]),
                    "Tipo": linha["Tipo"],
                }
            )

    return contas

def salvar_arquivo():
    # Criar uma janela Tkinter
    root = tk.Tk()
    root.withdraw()

    caminho_arquivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")]
    )

    if caminho_arquivo:
        pass
    else:
        # Obter o caminho da pasta de downloads
        downloads_folder = os.path.expanduser("~/Downloads")

        # Criar o caminho completo para o arquivo
        caminho_arquivo = os.path.join(downloads_folder, "calendario_anual.xlsx")

    return caminho_arquivo



def deleta_csv(caminho_arquivo):
    try:
    # Exclui o arquivo
        os.remove(caminho_arquivo)
        
    except OSError as e:
        print(f"Erro ao excluir o arquivo: {e}")

 
if __name__ == "__main__":


    print()
    barra_de_progresso("Iniciando o programa...")
    
    time.sleep(2)
    while True:
        arquivo_preenchido = input('\nA pasta de trabalho "Contas Basicas" de Recebimentos e Despesa ja esta preenchida (s/n): ').lower()
        
        if arquivo_preenchido == "s":
            break
        elif arquivo_preenchido == "n":
            print("\nPreencha os arquivos e reinicie o programa\n")
            time.sleep(2)    
            print("Programa encerrado")
            sys.exit()
        else:
            print("\nOpcao invalida\n")
            continue  
    
    print()
    
    excel_para_csv()
    
    verificar_vencimentos()
    
    main()
    
    print("\nDeletando arquivos CSV criados...\n")
    
    print("Deletando CSV de Recebimentos...")
    deleta_csv(arquivo_de_entradas)
    barra_de_progresso("Progresso")
    
    print("\nDeletando CSV de Despesas...")
    deleta_csv(arquivo_de_despesas)
    barra_de_progresso("Progresso")
    
    time.sleep(2)
    
    print()
    barra_de_progresso("Encerrando programa")
    time.sleep(0.5)
    print("\nPrograma encerrado com sucesso")
    
    